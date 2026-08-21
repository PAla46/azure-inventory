#!/usr/bin/env python3
"""One-shot Azure estate inventory exporter.

Queries every reachable resource via Azure Resource Graph, enriches it with
governance / networking / security / operations attributes derived from
resource properties and tags, and writes a flat CSV and/or XLSX file.

Authentication uses DefaultAzureCredential (az login, environment variables,
managed identity, VS Code credentials, ...).

Usage:
    az login
    pip install -r requirements.txt
    python azure_inventory.py --format both
"""

import argparse
import base64
import csv
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone

N_A = "N/A"

COLUMNS = [
    "Azure Resource ID",
    "Resource Name",
    "Resource Type",
    "Resource Category",
    "Parent Resource ID",
    "Tenant",
    "Management Group",
    "Subscription ID",
    "Subscription Name",
    "Resource Group",
    "Region",
    "Availability Zone",
    "SKU / Service Tier",
    "Resource State",
    "Associated Application ID",
    "Associated Application Name",
    "Business Unit / Function",
    "Business Owner",
    "Technical Owner",
    "Support Group",
    "Environment",
    "Criticality",
    "Data Classification",
    "Customer-Facing",
    "Regulatory Relevance",
    "Lifecycle Status",
    "Publicly Accessible",
    "Virtual Network / Subnet",
    "NSG / Network Control",
    "Managed Identity",
    "Tag Compliance Status",
    "Policy Compliance Status",
    "Encryption Status",
    "Logging and Monitoring Status",
    "Backup / Recovery Status",
    "Inventory Last Refreshed",
]

TAG_FIELDS = [
    "Associated Application ID",
    "Associated Application Name",
    "Business Unit / Function",
    "Business Owner",
    "Technical Owner",
    "Support Group",
    "Environment",
    "Criticality",
    "Data Classification",
    "Customer-Facing",
    "Regulatory Relevance",
    "Lifecycle Status",
]

# Tag keys are normalized (lowercase, separators stripped) before matching,
# e.g. tag "business-owner" matches pattern ^businessowner$. First match wins.
TAG_PATTERNS = {
    "Associated Application ID": [
        r"^applicationid$", r"^associatedapplicationid$", r"^appid$",
        r"^appidentification$", r"^appcode$", r"^applicationguid$",
        r"^moniker$", r"^appregistrationid$",
    ],
    "Associated Application Name": [
        r"^applicationname$", r"^associatedapplicationname$", r"^appname$",
        r"^application$", r"^app$", r"^servicename$", r"^productname$",
        r"^workloadname$",
    ],
    "Business Unit / Function": [
        r"^businessunit$", r"^bu$", r"^department$", r"^division$",
        r"^costcenter$", r"^costcentre$", r"^function$", r"^businessfunction$",
        r"^orgunit$", r"^organisationunit$", r"^organizationunit$",
    ],
    "Business Owner": [
        r"^businessowner$", r"^businesscontact$", r"^budgetowner$",
        r"^owner$", r"^ownedby$", r"^requestor$", r"^requestedby$",
    ],
    "Technical Owner": [
        r"^technicalowner$", r"^techowner$", r"^itowner$", r"^itcontact$",
        r"^technicalcontact$", r"^operatedby$", r"^devowner$", r"^sysadmin$",
        r"^administrator$", r"^admin$", r"^engineer$",
    ],
    "Support Group": [
        r"^supportgroup$", r"^supportteam$", r"^supportdl$", r"^support$",
        r"^team$", r"^distributionlist$", r"^dl$", r"^mail$", r"^email$",
        r"^contactemail$", r"^escalation$",
    ],
    "Environment": [
        r"^environment$", r"^environmenttype$", r"^env$", r"^envtype$",
        r"^stage$", r"^deploymentenvironment$",
    ],
    "Criticality": [
        r"^criticality$", r"^businesscriticality$", r"^importance$",
        r"^priority$", r"^tier$", r"^servicelevel$", r"^sla$",
        r"^impact$", r"^businessimpact$",
    ],
    "Data Classification": [
        r"^dataclassification$", r"^classification$", r"^sensitivity$",
        r"^confidentiality$", r"^datalabel$", r"^privacy$", r"^pii$",
    ],
    "Customer-Facing": [
        r"^customerfacing$", r"^customerfacingworkload$", r"^externalfacing$",
        r"^internetfacing$", r"^publicfacing$", r"^exposure$",
    ],
    "Regulatory Relevance": [
        r"^regulatoryrelevance$", r"^regulatory$", r"^regulated$",
        r"^compliance$", r"^compliancerelevant$", r"^gdpr$", r"^pci$",
        r"^pcidss$", r"^hipaa$", r"^sox$", r"^iso27001$", r"^nist$",
    ],
    "Lifecycle Status": [
        r"^lifecyclestatus$", r"^lifecycle$", r"^lifecyclestate$",
        r"^decommissiondate$", r"^retirementdate$", r"^endoflife$", r"^eol$",
        r"^sunsetting$", r"^status$",
    ],
}

REQUIRED_TAGS_DEFAULT = ["Environment", "Business Owner", "Technical Owner"]

CATEGORY_MAP = {
    "microsoft.compute": "Compute",
    "microsoft.storage": "Storage",
    "microsoft.network": "Networking",
    "microsoft.sql": "Databases",
    "microsoft.dbformysql": "Databases",
    "microsoft.dbforpostgresql": "Databases",
    "microsoft.dbformariadb": "Databases",
    "microsoft.documentdb": "Databases",
    "microsoft.cache": "Databases",
    "microsoft.web": "App Services",
    "microsoft.containerservice": "Containers",
    "microsoft.containerregistry": "Containers",
    "microsoft.app": "Containers",
    "microsoft.keyvault": "Security",
    "microsoft.security": "Security",
    "microsoft.securityinsights": "Security",
    "microsoft.customerlockbox": "Security",
    "microsoft.managedidentity": "Identity",
    "microsoft.insights": "Monitoring & Management",
    "microsoft.operationalinsights": "Monitoring & Management",
    "microsoft.operationsmanagement": "Monitoring & Management",
    "microsoft.automation": "Monitoring & Management",
    "microsoft.recoveryservices": "Backup & DR",
    "microsoft.dataprotection": "Backup & DR",
    "microsoft.logic": "Integration",
    "microsoft.eventgrid": "Integration",
    "microsoft.servicebus": "Integration",
    "microsoft.relay": "Integration",
    "microsoft.notificationhubs": "Integration",
    "microsoft.apimanagement": "Integration",
    "microsoft.kusto": "Analytics",
    "microsoft.synapse": "Analytics",
    "microsoft.databricks": "Analytics",
    "microsoft.hdinsight": "Analytics",
    "microsoft.search": "Analytics",
    "microsoft.streamanalytics": "Analytics",
    "microsoft.cognitiveservices": "AI & ML",
    "microsoft.machinelearningservices": "AI & ML",
    "microsoft.botservice": "AI & ML",
    "microsoft.devices": "IoT",
    "microsoft.timeseriesinsights": "IoT",
    "microsoft.digitaltwins": "IoT",
    "microsoft.desktopvirtualization": "End User Computing",
    "microsoft.cdn": "Networking",
    "microsoft.frontdoor": "Networking",
    "microsoft.trafficmanager": "Networking",
    "microsoft.network/dnszones": "Networking",
    "microsoft.authorization": "Governance",
    "microsoft.policyinsights": "Governance",
    "microsoft.management": "Governance",
    "microsoft.resources": "Governance",
    "microsoft.migrate": "Governance",
    "microsoft.devtestlab": "DevOps",
    "microsoft.devcenter": "DevOps",
    "microsoft.batch": "Compute",
}

REGION_MAP = {
    "eastus": "East US", "eastus2": "East US 2", "westus": "West US",
    "westus2": "West US 2", "westus3": "West US 3", "centralus": "Central US",
    "northcentralus": "North Central US", "southcentralus": "South Central US",
    "westcentralus": "West Central US", "canadacentral": "Canada Central",
    "canadaeast": "Canada East", "brazilsouth": "Brazil South",
    "brazilsoutheast": "Brazil Southeast", "northeurope": "North Europe",
    "westeurope": "West Europe", "uksouth": "UK South", "ukwest": "UK West",
    "francecentral": "France Central", "francesouth": "France South",
    "germanywestcentral": "Germany West Central", "germanynorth": "Germany North",
    "switzerlandnorth": "Switzerland North", "switzerlandwest": "Switzerland West",
    "norwayeast": "Norway East", "norwaywest": "Norway West",
    "swedencentral": "Sweden Central", "polandcentral": "Poland Central",
    "italynorth": "Italy North", "spaincentral": "Spain Central",
    "austriacentral": "Austria Central", "uaenorth": "UAE North",
    "uaecentral": "UAE Central", "israelcentral": "Israel Central",
    "qatarcentral": "Qatar Central", "southafricanorth": "South Africa North",
    "southafricawest": "South Africa West", "centralindia": "Central India",
    "southindia": "South India", "westindia": "West India",
    "japaneast": "Japan East", "japanwest": "Japan West",
    "koreacentral": "Korea Central", "koreasouth": "Korea South",
    "southeastasia": "Southeast Asia", "eastasia": "East Asia",
    "australiaeast": "Australia East", "australiacentral": "Australia Central",
    "australiacentral2": "Australia Central 2", "australiasoutheast": "Australia Southeast",
    "mexicocentral": "Mexico Central", "global": "Global",
}

ENV_HINTS = [
    (r"\b(prod|production|prd)\b", "Production"),
    (r"\b(dev|development|dvl)\b", "Development"),
    (r"\b(uat|stg|stage|staging|preprod|preprod?uction)\b", "Staging/UAT"),
    (r"\b(qa|test|tst)\b", "Test/QA"),
    (r"\b(sbx|sandbox)\b", "Sandbox"),
    (r"\b(demo|poc|prototype)\b", "PoC/Demo"),
]

SUBNET_RE = re.compile(
    r"/subscriptions/[^\"\s]+/resourceGroups/[^\"\s]+"
    r"/providers/Microsoft\.Network/virtualNetworks/[^\"\s/]+"
    r"(?:/subnets/[^\"\s/]+)?",
    re.I,
)
NSG_RE = re.compile(
    r"/subscriptions/[^\"\s]+/resourceGroups/[^\"\s]+"
    r"/providers/Microsoft\.Network/networkSecurityGroups/[^\"\s/]+",
    re.I,
)
SOURCE_RESOURCE_ID_RE = re.compile(r'"sourceResourceId"\s*:\s*"([^"]+)"', re.I)


def norm_key(key):
    return re.sub(r"[^a-z0-9]", "", str(key or "").lower())


def tag_lookup(tags, patterns):
    if not tags:
        return None
    for key in sorted(tags):
        nk = norm_key(key)
        if any(re.fullmatch(p, nk) for p in patterns):
            value = str(tags[key]).strip()
            return value or None
    return None


def infer_environment(text):
    low = f" {text.lower()} "
    for pattern, label in ENV_HINTS:
        if re.search(pattern, low):
            return label
    return None


def parent_resource_id(rid):
    parts = (rid or "").strip("/").split("/")
    indices = [i for i, seg in enumerate(parts) if seg.lower() == "providers"]
    if not indices:
        return ""
    last = indices[-1]
    if len(indices) > 1:
        return "/" + "/".join(parts[:last])
    if len(parts) - last - 1 > 3:
        return "/" + "/".join(parts[:-2])
    return ""


def resource_category(rtype):
    ns = (rtype or "").split("/")[0].lower()
    if ns == "microsoft.network" and len((rtype or "").split("/")) > 1:
        leaf = rtype.split("/")[1].lower()
        if leaf in ("dnszones", "privatednszones", "trafficmanagerprofiles"):
            return "Networking"
    return CATEGORY_MAP.get(ns) or ns.replace("microsoft.", "").title()


def pretty_region(slug):
    if not slug:
        return N_A
    return REGION_MAP.get(slug.lower()) or re.sub(r"(\d+)$", r" \1", slug).title()


def managed_identity_str(identity):
    itype = (identity or {}).get("type")
    if not itype:
        return "None"
    mapping = {
        "SystemAssigned": "System-assigned",
        "UserAssigned": "User-assigned",
        "SystemAssigned, UserAssigned": "System + User-assigned",
    }
    return mapping.get(itype, str(itype))


def sku_str(row, props):
    sku = row.get("sku") or props.get("sku") or {}
    if isinstance(sku, dict):
        name, tier = sku.get("name"), sku.get("tier")
        if name and tier and str(name) != str(tier):
            return f"{tier} / {name}"
        if name:
            return str(name)
        if tier:
            return str(tier)
    elif sku:
        return str(sku)
    return N_A


def resource_state(props):
    for key in ("provisioningState", "state", "status"):
        value = props.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    power = props.get("powerState")
    if isinstance(power, dict) and power.get("code"):
        return str(power["code"]).split("/")[-1]
    return N_A


def encryption_status(rtype, props, props_json_lower):
    tl = (rtype or "").lower()
    if tl.startswith("microsoft.storage/storageaccounts"):
        enc = props.get("encryption") or {}
        services = enc.get("services") or {}
        enabled = any(
            isinstance(services.get(s), dict) and services[s].get("enabled")
            for s in ("blob", "file", "queue", "table")
        )
        if not enabled:
            return "Disabled"
        source = str(enc.get("keySource") or "")
        return "Enabled (CMK)" if "keyvault" in source.lower() else "Enabled (Microsoft-managed)"
    if tl.startswith(("microsoft.compute/disks", "microsoft.compute/snapshots")):
        etype = str((props.get("encryption") or {}).get("type") or "")
        if "CustomerKey" in etype:
            return "Enabled (CMK)"
        if etype:
            return "Enabled (Platform-managed)"
        return N_A
    if tl.startswith("microsoft.keyvault"):
        return "Enabled (Platform-managed)"
    if "diskencryptionsetid" in props_json_lower or "keyvaulturi" in props_json_lower \
            or "customermanagedkey" in props_json_lower:
        return "Enabled (CMK referenced)"
    if '"encryption"' in props_json_lower:
        return "Enabled (review properties)"
    return N_A


def public_access(row, props, ctx):
    t = (row.get("type") or "").lower()
    rid = row["id"]
    if t == "microsoft.network/publicipaddresses":
        return "Yes (unattached)" if rid not in ctx["pip_attached"] else "Yes"
    if t == "microsoft.network/networkinterfaces":
        return "Yes (public IP)" if ctx["nic_to_pips"].get(rid) else "No"
    if t == "microsoft.compute/virtualmachines":
        for nic in ctx["vm_to_nics"].get(rid, []):
            if ctx["nic_to_pips"].get(nic):
                return "Yes (public IP)"
        return "No"
    if t in ("microsoft.network/loadbalancers", "microsoft.network/applicationgateways"):
        return "Yes (public frontend)" if ctx["lb_to_pips"].get(rid) else "No"
    if t == "microsoft.storage/storageaccounts":
        if str(props.get("publicNetworkAccess", "Enabled")).lower() == "disabled":
            return "No (public access disabled)"
        if str(props.get("allowBlobPublicAccess", "")).lower() == "true":
            return "Yes (blob public access)"
        action = str(((props.get("networkAcls") or {}).get("defaultAction")) or "Allow").lower()
        return "Yes" if action == "allow" else "Restricted (firewall/VNet)"
    if t.startswith("microsoft.web/sites") or t.startswith("microsoft.web/functions"):
        if str(props.get("publicNetworkAccess", "Enabled")).lower() == "disabled":
            return "No (public access disabled)"
        return "Yes"
    if t == "microsoft.keyvault/vaults":
        if str(props.get("publicNetworkAccess", "Enabled")).lower() == "disabled":
            return "No (public access disabled)"
        action = str(((props.get("networkAcls") or {}).get("defaultAction")) or "Allow").lower()
        return "Yes" if action == "allow" else "Restricted (firewall/VNet)"
    if t in ("microsoft.network/privateendpoints", "microsoft.network/privatednszones"):
        return "No (private)"
    return N_A


def vnet_subnet(row, props_json, ctx):
    t = (row.get("type") or "").lower()
    rid = row["id"]
    refs = []
    if t == "microsoft.compute/virtualmachines":
        for nic in ctx["vm_to_nics"].get(rid, []):
            subnet = ctx["nic_to_subnet"].get(nic)
            if subnet:
                refs.append(subnet)
    else:
        refs = [m.group(0) for m in SUBNET_RE.finditer(props_json)]
        if t == "microsoft.network/virtualnetworks" and rid not in refs:
            refs.insert(0, rid)
    seen, unique = set(), []
    for ref in refs:
        if ref not in seen:
            seen.add(ref)
            unique.append(ref)
    return "; ".join(unique[:3]) if unique else N_A


def nsg_control(row, props_json, ctx):
    t = (row.get("type") or "").lower()
    rid = row["id"]
    refs = []
    if t == "microsoft.network/networksecuritygroups":
        refs.append(f"{row['name']} (this NSG)")
    elif t == "microsoft.compute/virtualmachines":
        for nic in ctx["vm_to_nics"].get(rid, []):
            nsg = ctx["nic_to_nsg"].get(nic)
            if nsg:
                refs.append(nsg)
            subnet = ctx["nic_to_subnet"].get(nic)
            if subnet and subnet in ctx["subnet_to_nsg"]:
                refs.append(ctx["subnet_to_nsg"][subnet])
    else:
        refs.extend(m.group(0) for m in NSG_RE.finditer(props_json))
        for m in SUBNET_RE.finditer(props_json):
            sid = m.group(0)
            if "/subnets/" in sid and sid in ctx["subnet_to_nsg"]:
                refs.append(ctx["subnet_to_nsg"][sid])
    seen, unique = set(), []
    for ref in refs:
        if ref not in seen:
            seen.add(ref)
            unique.append(ref)
    return "; ".join(unique[:3]) if unique else N_A


def tag_compliance(tags, required_tags):
    if not required_tags:
        return N_A
    keys = {norm_key(k) for k in (tags or {})}
    missing = []
    for req in required_tags:
        patterns = TAG_PATTERNS.get(req)
        if patterns:
            found = any(re.fullmatch(p, nk) for nk in keys for p in patterns)
        else:
            found = norm_key(req) in keys
        if not found:
            missing.append(req)
    if not missing:
        return "Compliant"
    if len(missing) == len(required_tags):
        return "Non-compliant (no required tags)"
    return f"Partially compliant (missing: {', '.join(missing)})"


def build_network_ctx(rows):
    ctx = {
        "pip_attached": set(),
        "nic_to_pips": defaultdict(list),
        "nic_to_subnet": {},
        "nic_to_nsg": {},
        "subnet_to_nsg": {},
        "vm_to_nics": defaultdict(list),
        "lb_to_pips": defaultdict(list),
    }
    for row in rows:
        t = (row.get("type") or "").lower()
        rid = row["id"]
        props = row.get("properties") or {}
        try:
            if t == "microsoft.network/publicipaddresses":
                if '"ipconfiguration"' in json.dumps(props).lower():
                    ctx["pip_attached"].add(rid)
            elif t == "microsoft.network/networkinterfaces":
                for ipc in props.get("ipConfigurations") or []:
                    ipc_props = ipc.get("properties") or {}
                    pip = (ipc_props.get("publicIPAddress") or {}).get("id")
                    subnet = (ipc_props.get("subnet") or {}).get("id")
                    if pip:
                        ctx["nic_to_pips"][rid].append(pip)
                    if subnet:
                        ctx["nic_to_subnet"][rid] = subnet
                nsg = (props.get("networkSecurityGroup") or {}).get("id")
                if nsg:
                    ctx["nic_to_nsg"][rid] = nsg
            elif t == "microsoft.network/virtualnetworks":
                for sn in props.get("subnets") or []:
                    sn_props = sn.get("properties") or {}
                    nsg = (sn_props.get("networkSecurityGroup") or {}).get("id")
                    if sn.get("id") and nsg:
                        ctx["subnet_to_nsg"][sn["id"]] = nsg
            elif t == "microsoft.compute/virtualmachines":
                ids = [
                    ni.get("id")
                    for ni in ((props.get("networkProfile") or {}).get("networkInterfaces") or [])
                    if ni.get("id")
                ]
                ctx["vm_to_nics"][rid] = ids
            elif t in ("microsoft.network/loadbalancers", "microsoft.network/applicationgateways"):
                for fe in props.get("frontendIPConfigurations") or []:
                    pip = ((fe.get("properties") or {}).get("publicIPAddress") or {}).get("id")
                    if pip:
                        ctx["lb_to_pips"][rid].append(pip)
        except (AttributeError, TypeError):
            continue
    return ctx


def enrich(row, ctx):
    rid = row["id"]
    rtype = row.get("type") or ""
    props = row.get("properties") or {}
    tags = row.get("tags") or {}
    props_json = json.dumps(props, default=str)
    subscription_id = row.get("subscriptionId") or ""
    rg = row.get("resourceGroup") or ""
    name = row.get("name") or ""

    values = {}
    for field in TAG_FIELDS:
        values[field] = tag_lookup(tags, TAG_PATTERNS[field])

    if not values["Environment"]:
        values["Environment"] = infer_environment(f"{rg} {name}")

    lifecycle = values["Lifecycle Status"]
    if not lifecycle:
        state = resource_state(props)
        lifecycle = "Decommissioning" if state.lower() in ("deleting", "deleted") else None
    values["Lifecycle Status"] = lifecycle

    policy_states = ctx["policy"].get(rid.lower(), set())
    if "noncompliant" in policy_states:
        policy_status = "Non-Compliant"
    elif "compliant" in policy_states:
        policy_status = "Compliant"
    else:
        policy_status = N_A

    backup_status = N_A
    if rid.lower() in ctx["backup_ids"]:
        backup_status = "Protected (Azure Backup)"
    elif (rtype or "").lower() == "microsoft.recoveryservices/vaults":
        backup_status = "Recovery Services vault"

    record = {
        "Azure Resource ID": rid,
        "Resource Name": name,
        "Resource Type": rtype,
        "Resource Category": resource_category(rtype),
        "Parent Resource ID": parent_resource_id(rid) or N_A,
        "Tenant": ctx["tenant"],
        "Management Group": ctx["sub_mg"].get(subscription_id, N_A),
        "Subscription ID": subscription_id or N_A,
        "Subscription Name": ctx["sub_names"].get(subscription_id, N_A),
        "Resource Group": rg or N_A,
        "Region": pretty_region(row.get("location")),
        "Availability Zone": "; ".join(row.get("zones") or []) or N_A,
        "SKU / Service Tier": sku_str(row, props),
        "Resource State": resource_state(props),
        "Associated Application ID": values["Associated Application ID"] or N_A,
        "Associated Application Name": values["Associated Application Name"] or N_A,
        "Business Unit / Function": values["Business Unit / Function"] or N_A,
        "Business Owner": values["Business Owner"] or N_A,
        "Technical Owner": values["Technical Owner"] or N_A,
        "Support Group": values["Support Group"] or N_A,
        "Environment": values["Environment"] or N_A,
        "Criticality": values["Criticality"] or N_A,
        "Data Classification": values["Data Classification"] or N_A,
        "Customer-Facing": values["Customer-Facing"] or N_A,
        "Regulatory Relevance": values["Regulatory Relevance"] or N_A,
        "Lifecycle Status": values["Lifecycle Status"] or N_A,
        "Publicly Accessible": public_access(row, props, ctx),
        "Virtual Network / Subnet": vnet_subnet(row, props_json, ctx),
        "NSG / Network Control": nsg_control(row, props_json, ctx),
        "Managed Identity": managed_identity_str(row.get("identity")),
        "Tag Compliance Status": tag_compliance(tags, ctx["required_tags"]),
        "Policy Compliance Status": policy_status,
        "Encryption Status": encryption_status(rtype, props, props_json.lower()),
        "Logging and Monitoring Status": (
            "Diagnostic settings enabled" if rid.lower() in ctx["diag_parents"] else "Not detected"
        ),
        "Backup / Recovery Status": backup_status,
        "Inventory Last Refreshed": ctx["refreshed_at"],
    }
    return [record[c] for c in COLUMNS]


def get_credential():
    from azure.identity import DefaultAzureCredential

    return DefaultAzureCredential()


def get_arg_client(credential):
    from azure.mgmt.resourcegraph import ResourceGraphClient

    return ResourceGraphClient(credential)


def run_arg_query(client, query, subscriptions=None, page_size=200):
    from azure.core.exceptions import HttpResponseError
    from azure.mgmt.resourcegraph.models import QueryRequest, QueryRequestOptions

    rows, skip_token = [], None
    while True:
        options = QueryRequestOptions(result_format="objectArray", top=page_size)
        if skip_token:
            options.skip_token = skip_token
        request = QueryRequest(subscriptions=subscriptions, query=query, options=options)
        try:
            response = client.resources(request)
        except HttpResponseError as exc:
            message = str(exc).lower()
            if page_size > 20 and ("payload" in message or "too large" in message or "429" in message):
                page_size = max(20, page_size // 2)
                print(f"  Reducing page size to {page_size} and retrying...", file=sys.stderr)
                continue
            raise
        data = response.data or []
        rows.extend(data)
        skip_token = getattr(response, "skip_token", None)
        print(f"  fetched {len(rows)} rows...", file=sys.stderr)
        if not skip_token or not data:
            break
    return rows


def resolve_tenant(credential, rows):
    for row in rows:
        tid = row.get("tenantId")
        if tid:
            return tid
    try:
        token = credential.get_token("https://management.azure.com/.default").token
        payload = token.split(".")[1]
        payload += "=" * (-len(payload) % 4)
        claims = json.loads(base64.urlsafe_b64decode(payload))
        return claims.get("tid") or N_A
    except Exception:
        return N_A


def fetch_subscription_metadata(client, subscriptions):
    query = (
        "resourcecontainers | where type == 'microsoft.resources/subscriptions' "
        "| project subscriptionId, name, mgChain = tostring(properties.managementGroupAncestorChain)"
    )
    names, mg_map = {}, {}
    for row in run_arg_query(client, query, subscriptions):
        sub_id = row.get("subscriptionId")
        names[sub_id] = row.get("name") or N_A
        try:
            chain = json.loads(row.get("mgChain") or "[]")
        except (TypeError, json.JSONDecodeError):
            chain = []
        labels = []
        for entry in chain:
            label = (entry or {}).get("displayName") or (entry or {}).get("name")
            if label and label not in labels:
                labels.append(label)
        mg_map[sub_id] = " > ".join(labels) if labels else N_A
    return names, mg_map


def fetch_diagnostic_parent_ids(client, subscriptions):
    query = "resources | where type =~ 'microsoft.insights/diagnosticsettings' | project id"
    parents = set()
    for row in run_arg_query(client, query, subscriptions):
        parent = re.sub(
            r"/providers/microsoft\.insights/diagnosticsettings/[^/]+$",
            "",
            row.get("id", ""),
            flags=re.I,
        )
        if parent:
            parents.add(parent.lower())
    return parents


def fetch_backup_source_ids(client, subscriptions):
    query = (
        "resources | where type =~ 'microsoft.recoveryservices/vaults/backupprotecteditems' "
        "| project id, properties"
    )
    ids = set()
    for row in run_arg_query(client, query, subscriptions):
        props = row.get("properties") or {}
        blob = json.dumps(props, default=str)
        for match in SOURCE_RESOURCE_ID_RE.findall(blob):
            ids.add(match.lower())
        vm_id = (props.get("virtualMachineId") or props.get("sourceResourceId") or "")
        if vm_id:
            ids.add(str(vm_id).lower())
    return ids


def fetch_policy_compliance(client, subscriptions):
    query = (
        "policyresources | where type =~ 'microsoft.policyinsights/policystates' "
        "| extend rid = tostring(properties.resourceId), cs = tostring(properties.complianceState) "
        "| summarize states = make_set(cs) by rid"
    )
    states_by_resource = {}
    for row in run_arg_query(client, query, subscriptions):
        rid = (row.get("rid") or "").lower()
        states = {str(s).lower() for s in (row.get("states") or [])}
        if rid:
            states_by_resource[rid] = states
    return states_by_resource


def write_csv(path, records):
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(COLUMNS)
        writer.writerows(records)


def write_xlsx(path, records):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Azure Inventory"
    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    def as_cell(value):
        text = "" if value is None else str(value)
        return text[:32000]

    for record in records:
        sheet.append([as_cell(v) for v in record])

    widths = []
    for idx, column in enumerate(COLUMNS, start=1):
        longest = len(column)
        for record in records[:500]:
            longest = max(longest, min(len(str(record[idx - 1])), 60))
        widths.append(min(longest + 2, 60))
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Export a complete Azure resource inventory to CSV/XLSX."
    )
    parser.add_argument("-o", "--output", help="Output file base path without extension")
    parser.add_argument(
        "-f", "--format", choices=["auto", "csv", "xlsx", "both"], default="auto",
        help="Output format (default: xlsx if openpyxl installed, else csv)",
    )
    parser.add_argument(
        "-s", "--subscriptions", default="",
        help="Comma-separated subscription IDs to scope the scan (default: all accessible)",
    )
    parser.add_argument(
        "-t", "--required-tags", default=",".join(REQUIRED_TAGS_DEFAULT),
        help=f"Comma-separated canonical tag names for compliance check "
             f"(default: {','.join(REQUIRED_TAGS_DEFAULT)}; empty string disables)",
    )
    parser.add_argument("--page-size", type=int, default=200, help="ARG page size (default 200)")
    parser.add_argument(
        "--fast", action="store_true",
        help="Skip fetching resource properties; VNet/NSG/public-access/encryption fields become N/A",
    )
    parser.add_argument("--skip-policy", action="store_true", help="Skip policy compliance lookup")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    started = time.time()

    subscriptions = [s.strip() for s in args.subscriptions.split(",") if s.strip()] or None
    required_tags = [t.strip() for t in args.required_tags.split(",") if t.strip()]

    print("Authenticating...", flush=True)
    credential = get_credential()
    client = get_arg_client(credential)

    projection = (
        "project id, name, type, kind, location, subscriptionId, resourceGroup, "
        "tags, sku, identity, zones, tenantId, properties"
    )
    if args.fast:
        projection = (
            "project id, name, type, kind, location, subscriptionId, resourceGroup, "
            "tags, sku, tenantId"
        )
    main_query = f"resources | {projection}"

    print("Querying Azure Resource Graph for all resources...", flush=True)
    rows = run_arg_query(client, main_query, subscriptions, args.page_size)
    if not rows:
        print("No resources found for the current credentials.", file=sys.stderr)
        return 1
    print(f"Retrieved {len(rows)} resources.", flush=True)

    print("Fetching subscription / management group metadata...", flush=True)
    sub_names, sub_mg = fetch_subscription_metadata(client, subscriptions)

    print("Fetching diagnostic settings...", flush=True)
    diag_parents = fetch_diagnostic_parent_ids(client, subscriptions)

    print("Fetching backup protected items...", flush=True)
    backup_ids = fetch_backup_source_ids(client, subscriptions)

    policy = {}
    if not args.skip_policy:
        print("Fetching policy compliance states...", flush=True)
        try:
            policy = fetch_policy_compliance(client, subscriptions)
        except Exception as exc:
            print(f"  Policy lookup failed ({exc}); continuing with N/A.", file=sys.stderr)

    refreshed_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ctx = {
        "sub_names": sub_names,
        "sub_mg": sub_mg,
        "diag_parents": diag_parents,
        "backup_ids": backup_ids,
        "policy": policy,
        "tenant": resolve_tenant(credential, rows),
        "required_tags": required_tags,
        "refreshed_at": refreshed_at,
    }

    print("Building network relationship maps...", flush=True)
    ctx.update(build_network_ctx(rows))

    print("Enriching resources...", flush=True)
    records = []
    for index, row in enumerate(rows, start=1):
        try:
            records.append(enrich(row, ctx))
        except Exception as exc:
            print(f"  Skipping {row.get('id')}: {exc}", file=sys.stderr)
        if index % 1000 == 0:
            print(f"  enriched {index}/{len(rows)}...", flush=True)

    records.sort(key=lambda r: (r[COLUMNS.index("Subscription Name")],
                                r[COLUMNS.index("Resource Group")],
                                r[COLUMNS.index("Resource Type")],
                                r[COLUMNS.index("Resource Name")]))

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = args.output or os.path.join(os.getcwd(), f"azure_inventory_{stamp}")
    fmt = args.format
    if fmt == "auto":
        try:
            import openpyxl  # noqa: F401
            fmt = "xlsx"
        except ImportError:
            fmt = "csv"

    outputs = []
    if fmt in ("csv", "both"):
        path = f"{base}.csv"
        write_csv(path, records)
        outputs.append(path)
    if fmt in ("xlsx", "both"):
        path = f"{base}.xlsx"
        write_xlsx(path, records)
        outputs.append(path)

    elapsed = time.time() - started
    print(f"\nDone: {len(records)} resources inventoried in {elapsed:.1f}s")
    for path in outputs:
        print(f"  wrote {path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nAborted.", file=sys.stderr)
        sys.exit(130)
